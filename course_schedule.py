#! /usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl as opxl
import datetime

# def main():
# 	nntime = datetime.datetime.now()
# 	nnweek = nntime.isoweekday()
# 	results = get_day_courses(nnweek+1, '/home/dash/Documents/dash_courses.xlsx')
# 	if isinstance(results, str):
# 		print(results)
# 	else:
# 		for result in results:
# 			print(result.value)

# 得到下一节课的内容
def get_next_course(nnweek, nntime, schedule_file):
	# 操作 Excel 表格
	wb = opxl.load_workbook(schedule_file)
	sheetnames = wb.get_sheet_names()
	st0 = wb.get_sheet_by_name(sheetnames[0])
	col0 = st0['A']
	# 判断下一节课处于一星期中的哪一天
	(nntime, nnweek) = cal_weekday(nnweek, nntime)
	# 判断现在下节课的具体时间
	line = cal_time(nntime, col0, schedule_file)
	# 判断表格是否为空
	while st0.cell(row=line, column=nnweek+1).value == None:
		# 依次增加格子，直到找到合适的为止
		if line < 9:
			line = line + 1
		else:
			nnweek = nnweek + 1
			line = 2
			if nnweek > 5:
				nnweek = 1
	course_name = st0.cell(row=line, column=nnweek+1).value
	return course_name

# 得到一天的课程内容
def get_day_courses(nnweek, schedule_file):
	# 先判断是否有课，没课则直接返回
	if nnweek>5 and nnweek<=7:
		return ("当天没课，好好规划时间")
	elif nnweek > 7:
		nnweek = 1
	# 操作 Excel 表格
	wb = opxl.load_workbook(schedule_file)
	sheetnames = wb.get_sheet_names()
	st0 = wb.get_sheet_by_name(sheetnames[0])
	# 得到该天在 Excel 中的所在列
	day_need = chr(nnweek+ord('A'))
	col_need = st0[day_need]
	return col_need

# 计算下节课是周几
def cal_weekday(nnweek, nntime):
	if moment_cmp(nntime.hour, nntime.minute, datetime.datetime.strptime("18:55", "%H:%M")):
		nnweek = nnweek + 1
		nntime = nntime.replace(hour=0, minute=0)
	if nnweek>5:
		nnweek = 1
	return (nntime, nnweek)

# 计算下节课排在一天中的第几节
def cal_time(nntime, col, wb_file):
	line = 2
	#course_data = xlrd.open_workbook(wb_file)
	for start_time in col:
		# 跳过初始的空白单元格
		if start_time.value == None:
			continue
		#start_time = datetime.time(*xlrd.xldate_as_tuple(start_time, course_data.datemode)[3:])
		if moment_cmp(nntime.hour, nntime.minute, start_time.value):
			line = line + 1
		else:
			break
	return line

# 比较时钟和分钟
def moment_cmp(nhour, nmin, basetime):
	interval_time = int(nhour)*60 + int(nmin) - int(basetime.hour)*60 - int(basetime.minute)
	if interval_time > 0:
		return True
	else:
		return False

# main()
